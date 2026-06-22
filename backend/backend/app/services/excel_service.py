import logging
import subprocess
import shutil
import tempfile
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

EXCEL_PATH = Path(__file__).resolve().parents[2] / "excel" / "work_order.xlsx"


def _recalculate_with_libreoffice(path: Path) -> None:
    """
    Use LibreOffice headless to force-evaluate all formulas and cache their
    results back into the workbook.  Output is written to a temp dir then
    moved in place (LibreOffice cannot overwrite the source file directly).

    Silently skips if LibreOffice is not installed.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        logger.warning(
            "LibreOffice not found on PATH – formula cells will read as None. "
            "Install with: sudo apt-get install libreoffice"
        )
        return

    with tempfile.TemporaryDirectory() as tmp_dir:
        cmd = [soffice, "--headless", "--convert-to", "xlsx",
               "--outdir", tmp_dir, str(path)]
        logger.info("Recalculating workbook with LibreOffice …")
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            logger.error("LibreOffice recalculation failed: %s", result.stderr)
            return
        recalculated = Path(tmp_dir) / path.name
        if recalculated.exists():
            shutil.copy2(str(recalculated), str(path))
            logger.info("Workbook replaced with recalculated version.")
        else:
            logger.error("LibreOffice output not found: %s", recalculated)


def write_inputs_and_read_outputs(
    inputs: Dict[str, float],
    input_mappings: Dict[str, str],
    output_mappings: Dict[str, str],
    sheet_name: str,
) -> Dict[str, Any]:
    """
    Write input values into the Excel workbook using the cell mappings,
    trigger a LibreOffice recalculation, then read back the calculated outputs.

    Args:
        inputs:          dict of field_name -> value supplied by the client
        input_mappings:  dict of field_name -> cell address (e.g. "D1")
        output_mappings: dict of label -> cell address (e.g. "TL" -> "E1")
        sheet_name:      name of the worksheet to target

    Returns:
        dict of label -> calculated value read from the workbook
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel workbook not found at {EXCEL_PATH}")

    # ── Write inputs ──────────────────────────────────────────────────────────
    logger.info("Loading workbook for writing: %s", EXCEL_PATH)
    wb = load_workbook(EXCEL_PATH)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    for field, cell_addr in input_mappings.items():
        value = inputs.get(field)
        if value is None:
            logger.warning("Input '%s' not provided; skipping %s", field, cell_addr)
            continue
        ws[cell_addr] = value
        logger.debug("  Wrote %s -> %s = %s", field, cell_addr, value)

    wb.save(EXCEL_PATH)
    logger.info("Workbook saved with inputs.")

    # ── Recalculate formulas via LibreOffice ──────────────────────────────────
    _recalculate_with_libreoffice(EXCEL_PATH)

    # ── Read outputs ──────────────────────────────────────────────────────────
    logger.info("Reading outputs (data_only mode).")
    wb_data = load_workbook(EXCEL_PATH, data_only=True)
    ws_data = wb_data[sheet_name]

    outputs: Dict[str, Any] = {}
    for label, cell_addr in output_mappings.items():
        raw = ws_data[cell_addr].value
        outputs[label] = raw
        logger.debug("  Read %s <- %s = %s", label, cell_addr, raw)

    return outputs
